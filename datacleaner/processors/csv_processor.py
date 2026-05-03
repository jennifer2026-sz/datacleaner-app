"""CSV / Excel processor."""

import csv
from pathlib import Path


def process_csv(filepath: str | Path) -> str:
    """Read CSV and return text representation for scanning.

    Returns the CSV content as a flat text with column headers,
    so the PII detectors can analyze all cell values.
    """
    rows = []
    encodings = ["utf-8", "latin-1", "cp1252", "gbk", "gb2312"]

    for enc in encodings:
        try:
            with open(filepath, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                if headers:
                    rows.append(" | ".join(headers))
                    rows.append("-" * 60)
                for row in reader:
                    rows.append(" | ".join(row))
            return "\n".join(rows)
        except (UnicodeDecodeError, UnicodeError):
            continue

    raise ValueError(f"Unable to decode CSV file: {filepath}")


def process_excel(filepath: str | Path) -> str:
    """Read Excel (.xlsx/.xls) and return text representation."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    all_sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [f"--- Sheet: {sheet_name} ---"]

        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                clean_row = [str(cell) if cell is not None else "" for cell in row]
                rows.append(" | ".join(clean_row))

        all_sheets.append("\n".join(rows))

    wb.close()
    return "\n\n".join(all_sheets)


def get_csv_columns(filepath: str | Path) -> list[str]:
    """Get column headers of a CSV file."""
    with open(filepath, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        return next(reader, [])
